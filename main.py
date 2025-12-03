import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import glob

def parse_sql_file(sql_file_path):
    """Parse SQL file and extract table structures"""
    with open(sql_file_path, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()

    tables = {}

    # Pattern to match CREATE TABLE statements
    create_table_pattern = r'CREATE TABLE `(\w+)`\s*\((.*?)\)\s*ENGINE'

    matches = re.finditer(create_table_pattern, content, re.DOTALL | re.IGNORECASE)
    print("matches", matches)
    for match in matches:
        table_name = match.group(1)
        table_definition = match.group(2)

        # Extract columns (skip constraints)
        columns = []
        lines = table_definition.split('\n')

        for line in lines:
            line = line.strip()

            # Skip empty lines, comments, and constraint definitions
            if not line or line.startswith('--') or line.startswith('/*'):
                continue
            if line.upper().startswith(('PRIMARY KEY', 'KEY', 'UNIQUE KEY', 'CONSTRAINT', 'FOREIGN KEY', 'INDEX')):
                continue

            # Extract column name and type
            column_match = re.match(r'`(\w+)`\s+(.+)', line)
            if column_match:
                column_name = column_match.group(1)
                column_type = column_match.group(2).strip()

                # Remove trailing comma if exists
                if column_type.endswith(','):
                    column_type = column_type[:-1].strip()

                # Clean up the type - more comprehensive removal of SQL metadata
                # Remove COLLATE clauses (any collation)
                column_type = re.sub(r'\s+COLLATE\s+\w+', '', column_type, flags=re.IGNORECASE)
                
                # Remove CHARACTER SET clauses
                column_type = re.sub(r'\s+CHARACTER SET\s+\w+', '', column_type, flags=re.IGNORECASE)
                
                # Remove constraints and metadata
                column_type = re.sub(r'\s+NOT NULL', '', column_type, flags=re.IGNORECASE)
                column_type = re.sub(r'\s+NULL', '', column_type, flags=re.IGNORECASE)
                column_type = re.sub(r'\s+AUTO_INCREMENT', '', column_type, flags=re.IGNORECASE)
                column_type = re.sub(r'\s+DEFAULT\s+.*?(?=\s+|$)', '', column_type, flags=re.IGNORECASE)
                column_type = re.sub(r'\s+COMMENT\s+.*?(?=\s+|$)', '', column_type, flags=re.IGNORECASE)
                column_type = re.sub(r'\s+ON UPDATE\s+.*?(?=\s+|$)', '', column_type, flags=re.IGNORECASE)
                column_type = re.sub(r'\s+UNIQUE', '', column_type, flags=re.IGNORECASE)
                column_type = re.sub(r'\s+PRIMARY', '', column_type, flags=re.IGNORECASE)
                
                # Clean up extra whitespace
                column_type = re.sub(r'\s+', ' ', column_type).strip()

                columns.append({
                    'name': column_name,
                    'type': column_type,
                    'description': ''
                })

        if columns:
            tables[table_name] = columns

    return tables

def create_excel_from_tables(tables, output_file):
    """Create Excel file with table structures"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, size=11)

    table_name_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    table_name_font = Font(bold=True, size=11)

    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_alignment = Alignment(horizontal='center', vertical='center')

    # Create a sheet for all tables
    ws = wb.create_sheet("Database Schema")

    current_row = 1

    for table_name, columns in tables.items():
        # Table name header
        ws.merge_cells(f'A{current_row}:C{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = table_name
        cell.fill = table_name_fill
        cell.font = table_name_font
        cell.alignment = center_alignment
        cell.border = border_thin

        # Apply border to merged cells
        for col in ['B', 'C']:
            ws[f'{col}{current_row}'].border = border_thin

        current_row += 1

        # Column headers
        headers = ['No', 'Type Data', 'Description']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border_thin

        current_row += 1

        # Column data
        for idx, column in enumerate(columns, 1):
            # Column name (No)
            cell = ws.cell(row=current_row, column=1)
            cell.value = column['name']
            cell.border = border_thin

            # Type Data - set as text to preserve commas in decimal(x,y) and enum values
            cell = ws.cell(row=current_row, column=2)
            cell.value = column['type']
            cell.number_format = '@'  # Text format to preserve commas
            cell.border = border_thin

            # Description
            cell = ws.cell(row=current_row, column=3)
            cell.value = column['description']
            cell.border = border_thin

            current_row += 1

        # Add empty row between tables
        current_row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50

    # Save workbook
    wb.save(output_file)
    print(f"✓ Created: {output_file}")

def main():
    """Main function to process all SQL files in the current directory"""
    # Get all SQL files in current directory
    sql_files = glob.glob("*.sql")

    if not sql_files:
        print("No SQL files found in the current directory.")
        return

    print(f"Found {len(sql_files)} SQL file(s)")
    print("-" * 50)

    for sql_file in sql_files:
        print(f"\nProcessing: {sql_file}")

        # Parse SQL file
        tables = parse_sql_file(sql_file)
        print(f"  Found {len(tables)} table(s)")

        # Create output filename
        base_name = os.path.splitext(sql_file)[0]
        output_file = f"{base_name}.xlsx"

        # Create Excel file
        create_excel_from_tables(tables, output_file)

    print("\n" + "=" * 50)
    print("All SQL files have been converted to Excel!")
    print("=" * 50)

if __name__ == "__main__":
    main()
